import sys
import pandas as pd
import numpy as np
import datetime
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

def main():
  # --- 1. CONFIGURACIÓN INICIAL Y ARGUMENTOS ---
  # if len(sys.argv) < 2:
    # print("❌ Error: Se requiere el nombre del archivo de scraping.")
    # sys.exit(1)

  # archivo_scraping = sys.argv[1]

  hoy_excel = datetime.datetime.now().strftime("%d-%m-%Y")
  hoy_excel_corto = datetime.datetime.now().strftime("%d-%m-%y")
  hoy_iso = datetime.datetime.now().strftime("%Y-%m-%d")

  exports = Path("exports")
  archivo_scraping = exports / f"{hoy_iso}.xlsx"

  print(f"🚀 Iniciando Procesamiento Maestro para: {archivo_scraping}")

  def formatear_precio(valor):
    try:
      if pd.isna(valor) or valor == 0 or valor == '0':
        return np.nan
      return float(valor)
    except:
      return np.nan

  def aplicar_limpieza_precios(df, columnas):
    df[columnas] = df[columnas].replace([0, 0.0, '0', '0,00'], np.nan)
    df[columnas] = df[columnas].fillna("")
    return df

  # Carga de la fuente principal
  try:
    df_fv_raw = pd.read_excel(archivo_scraping)
    df_fv_raw['id'] = pd.to_numeric(df_fv_raw['id'], errors='coerce').fillna(0).astype(int)
    df_fv_raw['price'] = df_fv_raw['price'].apply(formatear_precio)
    # Columnas I y J: Barcodes (se convierten a números para quitar la advertencia de texto)
    df_fv_raw['bar_code'] = pd.to_numeric(df_fv_raw['bar_code'], errors='coerce')
    df_fv_raw['bar_code_2'] = pd.to_numeric(df_fv_raw['bar_code_2'], errors='coerce')
    print("✅ Fuente de precios cargada y formateada.")
  except Exception as e:
    print(f"❌ Error cargando el scraping: {e}")
    sys.exit(1)

  nombre_salida = f"F1 {hoy_excel}.xlsx"
  folder_path = Path("f1")
  folder_path.mkdir(parents=True, exist_ok=True)
  ruta_final = folder_path / nombre_salida

  try:
    with pd.ExcelWriter(ruta_final, engine='xlsxwriter') as writer:
      df_fv_raw.to_excel(writer, index=False, sheet_name='Sheet1')

      wb = writer.book
      ws = writer.sheets['Sheet1']

      # Definimos formatos para que Excel no use notación científica en los códigos
      format_precio = wb.add_format({'num_format': '#,##0.00'}) # Formato con decimales
      format_barcode = wb.add_format({'num_format': '0'})       # Formato número entero largo

      # Aplicamos los formatos a las columnas (A:A es 0, D:D es 3, I:I es 8, J:J es 9)
      ws.set_column('D:D', 15.27, format_precio)
      ws.set_column('I:J', 15.27, format_barcode)

    print(f"✅ Archivo guardado sin viñetas verdes en: {ruta_final}")
  except Exception as e:
    print(f"❌ Error al guardar con formato: {e}")

  # --- TAREA 1: MATRIZ CARGA F1 ---
  print(f"Iniciando Matriz Carga F1.xlsx")
  df_matriz = pd.read_excel('Matriz Carga F1.xlsx')
  col_j_original = df_matriz.columns[9]
  df_matriz = df_matriz.rename(columns={col_j_original: 'unidad'})
  df_matriz['sku FV'] = pd.to_numeric(df_matriz['sku FV'], errors='coerce').fillna(0).astype(int)

  res_carga = pd.merge(df_matriz, df_fv_raw[['id', 'price']], left_on='sku FV', right_on='id', how='left')

  # Lógica de unidades
  res_carga['F1_DE'] = res_carga['price']
  res_carga['F1_FG'] = res_carga['price']
  res_carga.loc[res_carga['unidad'] == 1, 'F1_DE'] = np.nan
  res_carga.loc[res_carga['unidad'] != 1, 'F1_FG'] = np.nan

  # Asignar a columnas oficiales
  res_carga['F1'] = res_carga['F1_DE']
  res_carga['Precio Final'] = res_carga['F1_DE']
  res_carga['F1E.1'] = res_carga['F1_FG']
  res_carga['Precio Final.1'] = res_carga['F1']

  # Limpieza y Estructura
  res_carga = aplicar_limpieza_precios(res_carga, ['F1', 'Precio Final', 'F1.1', 'Precio Final.1'])
  # res_carga['unidad'] = ""
  res_carga['Fecha'] = pd.to_datetime(hoy_excel, format="%d-%m-%Y")
  res_carga = res_carga.drop(columns=['id', 'price', 'F1_DE', 'F1_FG'])

  # Forzar nombres para Web
  res_carga.columns.values[5] = 'F1'
  res_carga.columns.values[6] = 'Precio Final'
  res_carga.columns.values[9] = ""

  # crear carpeta de destino
  folder_carga = Path("carga_f1")
  folder_carga.mkdir(parents=True, exist_ok=True)
  nombre_archivo = f"Carga F1 {hoy_excel_corto}.xlsx"
  ruta_final2 = folder_carga / nombre_archivo

  try:
    with pd.ExcelWriter(ruta_final2, engine='xlsxwriter', datetime_format='dd-mm-yyyy') as writer:
      res_carga.to_excel(writer, index=False, sheet_name='Sheet1')

      wb = writer.book
      ws = writer.sheets['Sheet1']

      # Formatos
      format_codigobarra = wb.add_format({'num_format': '0'}) # códigos
      format_precio = wb.add_format({'num_format': '#,##0.00'}) # número con decimales
      format_numero = wb.add_format({'num_format': '0'})
      # format_fecha = wb.add_format({'num_format': 'dd-mm-yyyy', 'align': 'right'})
      format_fecha = wb.add_format({'num_format': 'dd-mm-yyyy', 'align': 'right'})


      # Aplicar formatos
      ws.set_column('A:A', 14.18, format_codigobarra)
      ws.set_column('C:C', 40)
      ws.set_column('D:G', 11.82, format_precio)
      ws.set_column('H:H', None, format_numero)
      ws.set_column('I:I', 10, format_fecha)
      ws.set_column('M:M', None, format_numero)
      # ws.write('K1', res_carga.columns[10], format_numero)

    print(f"✅ Archivo guardado sin viñetas verdes en: {ruta_final2}")
  except Exception as e:
    print(f"❌ Error al guardar con formato: {e}")

  # --- TAREA 2: MATRIZ
  print(f"Iniciando Matriz Consumo (al 23-10-25).xlsx")
  def procesar_consumo(hoja):
    df = pd.read_excel('Matriz.xlsx', sheet_name=hoja, skiprows=2)
    df['Columna1'] = pd.to_numeric(df['Columna1'], errors='coerce').fillna(0).astype(int)
    df = pd.merge(df, df_fv_raw[['id', 'price']], left_on='Columna1', right_on='id', how='left')
    df['Farmavalue'] = df['price'].replace([0, 0.0, '0', '0,00'], np.nan).fillna("")
    return df.drop(columns=['Columna1', 'id', 'price'])

  df_final = procesar_consumo('F2')
  df_final = procesar_consumo('F3')

  # 1. DEFINIR Y CREAR LA CARPETA (Asegura que n8n siempre encuentre la ruta)
  folder_matriz = Path("matriz_consumo")
  folder_matriz.mkdir(parents=True, exist_ok=True)

  # Guardado con Estilo
  nombre_final_consumo = f"Precios {hoy_excel}.xlsx"
  ruta_final_consumo = folder_matriz / nombre_final_consumo

  with pd.ExcelWriter(ruta_final_consumo, engine='openpyxl') as writer:
    df_bomba_final.to_excel(writer, sheet_name='Bomba', startrow=2, index=False)
    df_fischel_final.to_excel(writer, sheet_name='Fischel', startrow=2, index=False)

    # Estilos de encabezado y bordes
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    highlight_fill = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))

    for sheet_name, title in [('Bomba', 'FLB:Reporte de códigos de barra'),
                                ('Fischel', 'Reporte de códigos de barra Fischel')]:
      ws = writer.sheets[sheet_name]
      ws['A1'] = title
      ws['A1'].font = Font(size=14, bold=True, color="1F4E78")

      # 1. Aplicar bordes a toda la tabla y estilo al encabezado
      for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
          cell.border = thin_border
          if cell.row == 3:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

      # 2. DEFINIR ANCHO PEQUEÑO PARA COLUMNA E (Descripción)
      # Ajustar el ancho a 20. Al no poner 'wrap_text', el texto se verá cortado
      # en la visualización de Excel si la celda de al lado tiene datos.
      ws.column_dimensions['A'].width = 14
      ws.column_dimensions['B'].width = 14
      ws.column_dimensions['C'].width = 14
      ws.column_dimensions['D'].width = 20
      ws.column_dimensions['E'].width = 40

      # 3. Auto-ajustar las demás columnas (A, B, C, D, F)
      for col in ['A', 'B', 'C', 'D', 'F']:
        max_length = 0
        for cell in ws[col]:
          try:
            if len(str(cell.value)) > max_length:
              max_length = len(str(cell.value))
          except: pass
        ws.column_dimensions[col].width = max_length + 2

      # 4. Color de destaque para Farma (Columna F)
      for cell in ws['F']:
        if cell.row > 3:
          cell.fill = highlight_fill
      ws.column_dimensions['A'].width = 14
      ws.column_dimensions['E'].width = 40

      # 5. Formatear Columna D como número entero (sin notación científica)
      # Empezamos desde la fila 4 para no afectar el encabezado (que está en la fila 3)
      for cell in ws['D']:
        if cell.row > 3:
        # "0" le dice a Excel que muestre el número completo sin decimales
          cell.number_format = '0'
      # Forzamos un ancho suficiente para que el código largo se vea bien
      ws.column_dimensions['D'].width = 25

  print(f"✅ Tarea 2 terminada: {nombre_final_consumo}")

  # --- TAREA 3: FARMA  ---
  print(f"Iniciando Matriz Extra")
  df_extra = pd.read_excel('Matriz.xlsx')
  df_extra['ID'] = pd.to_numeric(df_extra['ID FV'], errors='coerce').fillna(0).astype(int)
  df_extra_final = pd.merge(df_extra, df_fv_raw[['id', 'price']], left_on='ID FV', right_on='id', how='left')
  df_extra_final['Precio'] = df_extra_final['price'].replace([0, 0.0, '0', '0,00'], np.nan).fillna("")
  df_extra_final = df_extra_final.drop(columns=['ID FV', 'id', 'price'])

  # 1. DEFINIR Y CREAR LA CARPETA ESPECÍFICA
  folder_extra = Path("farma_extra")
  folder_extra.mkdir(parents=True, exist_ok=True)

  # Estilo
  nombre_extra = f"extra {hoy_excel}.xlsx"
  ruta_final_extra = folder_extra / nombre_extra

  with pd.ExcelWriter(ruta_final_extra, engine='openpyxl') as writer:
    df_extra_final.to_excel(writer, sheet_name='Hoja1', index=False)
    ws = writer.sheets['Hoja1']

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # Azul oscuro
    header_font = Font(color="FFFFFF", bold=True)
    highlight_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid") # Azul claro
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))

    # Aplicar estilos a la tabla (Ahora son 5 columnas: A a E)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5):
      for cell in row:
        cell.border = thin_border
        if cell.row == 1: # Encabezado
          cell.fill = header_fill
          cell.font = header_font
          cell.alignment = Alignment(horizontal="center")
        else:
          cell.alignment = Alignment(horizontal="left")

    # Definir anchos de columna (Ajustado a los requerimientos previos)
    ws.column_dimensions['A'].width = 18 # Código Barra
    ws.column_dimensions['B'].width = 15 # Código Dokka
    ws.column_dimensions['C'].width = 40 # Descripción FINAL
    ws.column_dimensions['D'].width = 15 # Precio FV
    ws.column_dimensions['E'].width = 15 # Formato

    # Resaltar la columna Precio FV (Columna D)
    for cell in ws['D']:
      if cell.row > 1:
        cell.fill = highlight_fill

    # 5. Formatear Columna D como número entero (sin notación científica)
    # Empezamos desde la fila 2 para no afectar el encabezado (que está en la fila 3)
    for cell in ws['A']:
      if cell.row > 1:
        # "0" le dice a Excel que muestre el número completo sin decimales
        cell.number_format = '0'
    # Forzamos un ancho suficiente para que el código largo se vea bien
    ws.column_dimensions['A'].width = 25

  print(f"✅ Tarea 3 terminada: {nombre_farma_extra}")

  print("✅ Todo procesado con éxito.")

if __name__ == "__main__":
  main()